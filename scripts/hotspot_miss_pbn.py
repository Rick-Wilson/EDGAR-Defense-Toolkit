#!/usr/bin/env python3
"""Generate a PBN file of hotspot miss boards from an EDGAR Defense workbook.

Extracts deal, auction, opening lead, and hotspot commentary from the
Hotspots Table sheet, then invokes pbn-to-pdf to produce a PDF.

Usage:
    python3 hotspot_miss_pbn.py <workbook.xlsx> [--pbn-to-pdf PATH]
"""

import argparse
import subprocess
import urllib.parse
from pathlib import Path

import openpyxl

PBN_TO_PDF = "/Applications/Bridge Utilities/pbn-to-pdf"

DEALER_MAP = {"1": "S", "2": "W", "3": "N", "4": "E"}
# BBO md dealer digit: 1=S, 2=W, 3=N, 4=E

VUL_MAP = {
    "o": "None",
    "n": "NS",
    "e": "EW",
    "b": "Both",
}

SUIT_NAMES = {"S": "Spades", "H": "Hearts", "D": "Diamonds", "C": "Clubs"}


def shorten_name(name):
    """Shorten 'FirstName_LastName' to 'FirstName...'."""
    if "_" in name:
        return name.split("_")[0] + "..."
    return name

# Opening lead hotspot categories
LEAD_CATEGORIES = {"Kxx_vsSuit", "Weird_OLs", "Ax+_Low"}

# Clockwise: N->E->S->W->N
NEXT_SEAT = {"N": "E", "E": "S", "S": "W", "W": "N"}


def parse_lin_url(url):
    """Parse a BBO handviewer LIN URL into deal components."""
    decoded = urllib.parse.unquote(url)

    def extract_field(tag):
        # Find |tag|value| or tag|value|
        marker = f"|{tag}|"
        idx = decoded.find(marker)
        if idx < 0:
            marker = f"={tag}|"
            idx = decoded.find(marker)
        if idx < 0:
            return None
        start = idx + len(marker)
        end = decoded.find("|", start)
        return decoded[start:end] if end >= 0 else decoded[start:]

    # Player names: pn|S,W,N,E|
    pn = extract_field("pn")
    players = pn.split(",") if pn else ["", "", "", ""]
    if len(players) < 4:
        players += [""] * (4 - len(players))

    # Deal: md|<dealer_digit><south_hand>,<west_hand>,<north_hand>,<east_opt>|
    md = extract_field("md")
    if not md:
        return None

    dealer_digit = md[0]
    dealer = DEALER_MAP.get(dealer_digit, "N")
    hands_str = md[1:]

    # Vulnerability: sv|<char>|
    sv = extract_field("sv")
    vul = VUL_MAP.get(sv, "None") if sv else "None"

    # Board number: ah|Board N|
    ah = extract_field("ah")
    board_num = ""
    if ah:
        board_num = ah.replace("Board ", "").replace("+", " ").strip()

    # Parse hands from md format: S-cards H-cards D-cards C-cards
    # md hands are in order: S, W, N (E is derived)
    hand_parts = hands_str.rstrip(",").split(",")

    def parse_bbo_hand(h):
        """Parse 'S23JH45KD9C2AQ' into PBN 'J32.K54.9.AQ2'"""
        suits = {"S": "", "H": "", "D": "", "C": ""}
        current = None
        for ch in h.upper():
            if ch in suits:
                current = ch
            elif current:
                suits[current] += ch
        return f"{suits['S']}.{suits['H']}.{suits['D']}.{suits['C']}"

    south = parse_bbo_hand(hand_parts[0]) if len(hand_parts) > 0 else "..."
    west = parse_bbo_hand(hand_parts[1]) if len(hand_parts) > 1 else "..."
    north = parse_bbo_hand(hand_parts[2]) if len(hand_parts) > 2 else "..."

    # Derive East hand from remaining cards
    if len(hand_parts) > 3 and hand_parts[3]:
        east = parse_bbo_hand(hand_parts[3])
    else:
        east = derive_east(north, south, west)

    # PBN deal format: "dealer:N_hand E_hand S_hand W_hand"
    deal_pbn = f"{dealer}:{north} {east} {south} {west}"

    # Auction: mb|bid| sequences
    auction = []
    pos = 0
    while True:
        mb_idx = decoded.find("|mb|", pos)
        if mb_idx < 0:
            break
        start = mb_idx + 4
        end = decoded.find("|", start)
        if end < 0:
            break
        bid = decoded[start:end]
        # Skip alert annotations that follow
        if decoded[end:end + 3] == "|an|":
            pass  # annotation follows, just skip it
        auction.append(normalize_bid(bid))
        pos = end  # don't skip the closing |, it's shared with the next tag

    # Opening lead: first pc|card| after auction
    lead = None
    pc_idx = decoded.find("|pc|")
    if pc_idx >= 0:
        start = pc_idx + 4
        end = decoded.find("|", start)
        if end >= 0:
            lead = decoded[start:end].upper()

    return {
        "players": players,  # [S, W, N, E]
        "dealer": dealer,
        "vul": vul,
        "deal": deal_pbn,
        "board": board_num,
        "auction": auction,
        "lead": lead,
    }


def normalize_bid(bid):
    """Normalize a BBO bid to PBN format."""
    bid = bid.upper()
    bid = bid.replace("!", "")  # strip alert marker
    if bid == "P":
        return "pass"
    if bid == "D":
        return "X"
    if bid == "R":
        return "XX"
    # Convert suit names: 1C, 2H, 3N, etc.
    bid = bid.replace("N", "NT") if len(bid) == 2 and bid[1] == "N" else bid
    return bid


def derive_east(north, south, west):
    """Derive East hand from the other three hands."""
    all_cards = {
        "S": set("AKQJT98765432"),
        "H": set("AKQJT98765432"),
        "D": set("AKQJT98765432"),
        "C": set("AKQJT98765432"),
    }

    for hand in [north, south, west]:
        suits = hand.split(".")
        suit_order = ["S", "H", "D", "C"]
        for i, cards in enumerate(suits):
            if i < 4:
                for c in cards.upper():
                    all_cards[suit_order[i]].discard(c)

    east_suits = []
    for suit in ["S", "H", "D", "C"]:
        east_suits.append("".join(sorted(all_cards[suit],
                                         key=lambda c: "AKQJT98765432".index(c))))
    return ".".join(east_suits)


def format_auction_pbn(auction, dealer):
    """Format auction as PBN lines (4 bids per line, starting at dealer)."""
    if not auction:
        return ""

    lines = []
    for i in range(0, len(auction), 4):
        chunk = auction[i:i + 4]
        lines.append(" ".join(chunk))

    return "\n".join(lines)


def format_lead(lead):
    """Format opening lead card as a PBN play tag value."""
    if not lead or len(lead) < 2:
        return None
    suit = lead[0]
    rank = lead[1:]
    return f"{suit}{rank}"


def main():
    parser = argparse.ArgumentParser(
        description="Generate PBN + PDF of hotspot miss boards."
    )
    parser.add_argument("workbook", type=Path,
                        help="Path to the EDGAR Defense .xlsx workbook")
    parser.add_argument("--pbn-to-pdf", type=Path, default=Path(PBN_TO_PDF),
                        help="Path to pbn-to-pdf binary")
    parser.add_argument("--boards-per-page", type=int, default=4,
                        help="Boards per page in PDF (default: 4)")
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Error: {args.workbook} not found")
        return

    output_dir = args.workbook.parent
    pbn_path = output_dir / "hotspot_misses.pbn"
    pdf_path = output_dir / "hotspot_misses.pdf"

    print(f"Loading {args.workbook.name}...")
    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)

    # Find Hotspots sheet
    ws = None
    for name in wb.sheetnames:
        if name.lower().startswith("hotspot"):
            ws = wb[name]
            break
    if ws is None:
        print("Error: No Hotspots sheet found")
        return

    # Load declarer mapping from Hand Records (BBO col -> Dec col)
    hr_ws = None
    for name in wb.sheetnames:
        if name.lower().startswith("hand records"):
            hr_ws = wb[name]
            break
    declarer_map = {}  # board_id -> declarer seat (N/S/E/W)
    if hr_ws:
        hr_headers = [cell.value for cell in hr_ws[1]]
        bbo_col = hr_headers.index("BBO") if "BBO" in hr_headers else None
        dec_col = hr_headers.index("Dec") if "Dec" in hr_headers else None
        if bbo_col is not None and dec_col is not None:
            for row in hr_ws.iter_rows(min_row=2, max_col=max(bbo_col, dec_col) + 1):
                bbo = row[bbo_col].value
                dec = row[dec_col].value
                if bbo is not None and dec:
                    try:
                        declarer_map[int(bbo)] = dec.strip().upper()
                    except (ValueError, TypeError):
                        pass
        print(f"  Loaded {len(declarer_map)} declarer mappings from Hand Records")

    # Headers: Hotspot ID, Class_Index, Link, Tinyurl, Board ID, Category,
    #          Subclass, Contract, Lead, Date, Player, Hit/Miss, Board, Vul, Vul_Side

    misses = []
    for row in ws.iter_rows(min_row=2, max_col=15):
        vals = [cell.value for cell in row]
        hit_miss = vals[11]
        if hit_miss != "Miss":
            continue

        hotspot_id = vals[0]
        lin_url = vals[3]  # Tinyurl column actually has full LIN URL in anon workbook
        board_id = vals[4]
        category = vals[5] or ""
        subclass = vals[6] or ""
        contract = vals[7] or ""
        lead_str = vals[8] or ""
        date = vals[9]
        player = vals[10] or ""

        if not lin_url or not isinstance(lin_url, str):
            continue
        if "handviewer" not in lin_url and "lin=" not in lin_url:
            continue

        parsed = parse_lin_url(lin_url)
        if parsed is None:
            print(f"  Warning: could not parse LIN URL for hotspot {hotspot_id}")
            continue

        misses.append({
            "hotspot_id": hotspot_id,
            "board_id": board_id,
            "category": category,
            "subclass": subclass,
            "contract": contract,
            "lead": lead_str,
            "date": date,
            "player": player,
            "parsed": parsed,
        })

    wb.close()
    print(f"  Found {len(misses)} hotspot misses with parseable LIN URLs")

    # Sort by hotspot ID
    misses.sort(key=lambda m: m["hotspot_id"] or 0)

    # Write PBN file
    with open(pbn_path, "w") as f:
        f.write("% PBN 2.1\n")
        f.write("% EXPORT\n")
        f.write("\n")

        for i, miss in enumerate(misses):
            p = miss["parsed"]
            board_label = str(i + 1)

            # Date string
            date_str = ""
            if miss["date"]:
                if hasattr(miss["date"], "strftime"):
                    date_str = miss["date"].strftime("%Y.%m.%d")
                else:
                    date_str = str(miss["date"])[:10].replace("-", ".")

            # Commentary: hotspot type + details
            commentary_parts = []
            if miss["category"]:
                label = miss["category"]
                if miss["subclass"]:
                    label += f" / {miss['subclass']}"
                commentary_parts.append(f"<b>Hotspot:</b> {label}")
            commentary_parts.append(
                f"<b>Contract:</b> {miss['contract']}  "
                f"<b>Lead:</b> {miss['lead']}  "
                f"<b>Player:</b> {miss['player']}"
            )
            if date_str:
                commentary_parts.append(f"<b>Date:</b> {date_str.replace('.', '-')}")
            commentary = "\n".join(commentary_parts)

            # Standard PBN tag order (matching bridge-parsers writer)
            f.write(f'[Event ""]\n')
            f.write(f'[Site ""]\n')
            f.write(f'[Date "{date_str}"]\n')
            f.write(f'[Board "{board_label}"]\n')
            f.write(f'[West "{shorten_name(p["players"][1])}"]\n')
            f.write(f'[North "{shorten_name(p["players"][2])}"]\n')
            f.write(f'[East "{shorten_name(p["players"][3])}"]\n')
            f.write(f'[South "{shorten_name(p["players"][0])}"]\n')
            f.write(f'[Dealer "{p["dealer"]}"]\n')
            f.write(f'[Vulnerable "{p["vul"]}"]\n')
            f.write(f'[Deal "{p["deal"]}"]\n')
            f.write(f'[Scoring ""]\n')
            f.write(f'[Declarer ""]\n')
            f.write(f'[Contract "{miss["contract"]}"]\n')
            f.write(f'[Result ""]\n')

            # Auction
            if p["auction"]:
                f.write(f'[Auction "{p["dealer"]}"]\n')
                f.write(format_auction_pbn(p["auction"], p["dealer"]) + "\n")

            # Opening lead for lead-related hotspot categories
            if miss["category"] in LEAD_CATEGORIES and p["lead"]:
                lead_card = format_lead(p["lead"])
                board_id = miss["board_id"]
                declarer = declarer_map.get(board_id) if board_id else None
                if lead_card and declarer and declarer in NEXT_SEAT:
                    leader_seat = NEXT_SEAT[declarer]
                    f.write(f'[Play "{leader_seat}"]\n')
                    f.write(f'{lead_card} - - -\n')

            f.write(f'{{{commentary}}}\n')
            f.write("\n")

    print(f"  Wrote {len(misses)} boards to {pbn_path}")

    # Run pbn-to-pdf
    if args.pbn_to_pdf.exists():
        cmd = [
            str(args.pbn_to_pdf),
            str(pbn_path),
            "-o", str(pdf_path),
            "-n", str(args.boards_per_page),
        ]
        print(f"  Running: {' '.join(cmd)}")
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            print(f"  PDF saved: {pdf_path}")
        else:
            print(f"  pbn-to-pdf failed: {result.stderr}")
    else:
        print(f"  pbn-to-pdf not found at {args.pbn_to_pdf}")
        print(f"  PBN file ready at {pbn_path}")


if __name__ == "__main__":
    main()
