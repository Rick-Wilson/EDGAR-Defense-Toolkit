#!/usr/bin/env python3
"""Generate packed timeline charts from an EDGAR Defense workbook.

Each hand gets its own dot, packed left-to-right within year bands.
Produces three PDFs in the same directory as the workbook:
1. All boards with ACBL disclosure boards in red
2. Same with larger dots
3. Same plus hotspot Miss boards in blue

Usage:
    python3 timeline_chart.py <workbook.xlsx>
"""

import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl


def load_data(workbook_path):
    """Load hand records, ACBL board IDs, and hotspot miss board IDs."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    # Find sheets by prefix (names may include case-specific suffixes)
    def find_sheet(prefix):
        for name in wb.sheetnames:
            if name.lower().startswith(prefix.lower()):
                return wb[name]
        return None

    # Hand Records sheet: Board ID (col A), Event name (col F), BBO (col G)
    ws = find_sheet("Hand Records")
    if ws is None:
        raise ValueError("No 'Hand Records' sheet found in workbook")

    boards = []  # list of (bbo_id, datetime)
    for row in ws.iter_rows(min_row=2, max_col=7):
        event = row[5].value     # Event name (datetime or string)
        bbo = row[6].value       # BBO board ID
        if event is None or bbo is None:
            continue
        if isinstance(event, str):
            try:
                event = datetime.strptime(event[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        try:
            boards.append((int(bbo), event))
        except (ValueError, TypeError):
            pass

    # ACBL Analysis Table: board IDs in col D
    acbl_ids = set()
    ws2 = find_sheet("ACBL Analysis")
    if ws2 is not None:
        for row in ws2.iter_rows(min_row=2, max_col=4):
            val = row[3].value
            if val is not None:
                try:
                    acbl_ids.add(int(val))
                except (ValueError, TypeError):
                    pass

    # Hotspots Table: Board ID (col E), Hit/Miss (col L)
    miss_ids = set()
    ws3 = find_sheet("Hotspots")
    if ws3 is not None:
        for row in ws3.iter_rows(min_row=2, max_col=12):
            board_id = row[4].value
            hit_miss = row[11].value
            if hit_miss == "Miss" and board_id is not None:
                try:
                    miss_ids.add(int(board_id))
                except (ValueError, TypeError):
                    pass

    wb.close()
    return boards, acbl_ids, miss_ids


def make_packed_plot(boards, acbl_ids, title, output_path,
                     dot_size=4, miss_ids=None):
    """Create a packed grid chart with one dot per hand, grouped by year.

    Hands are sorted by date within each year and packed left-to-right,
    wrapping to the next row. Years are separated by gaps on the Y axis.
    """
    # Group by year, sorted by date within each year
    by_year = defaultdict(list)
    for bbo_id, dt in boards:
        by_year[dt.year].append((dt, bbo_id))
    for year in by_year:
        by_year[year].sort()

    # Layout: pack into rows of COLS_PER_ROW
    cols_per_row = 150
    year_gap = 2  # rows of gap between years

    # Build (x, y, color) for each dot
    bg_x, bg_y = [], []
    acbl_x, acbl_y = [], []
    miss_x, miss_y = [], []

    years = sorted(by_year.keys())
    y_cursor = 0
    year_label_positions = []  # (y_pos, year_label)

    for year in years:
        hands = by_year[year]
        num_rows = (len(hands) + cols_per_row - 1) // cols_per_row
        year_label_positions.append((y_cursor + num_rows / 2, str(year)))

        for i, (dt, bbo_id) in enumerate(hands):
            x = i % cols_per_row
            row_in_year = i // cols_per_row
            y = y_cursor + row_in_year

            if bbo_id in acbl_ids:
                acbl_x.append(x)
                acbl_y.append(y)
            elif miss_ids and bbo_id in miss_ids:
                miss_x.append(x)
                miss_y.append(y)
            else:
                bg_x.append(x)
                bg_y.append(y)

        y_cursor += num_rows + year_gap

    total_rows = y_cursor - year_gap

    # Plot
    fig_height = 10
    fig, ax = plt.subplots(figsize=(16, fig_height))
    fig.patch.set_facecolor("#1a1a2e")
    ax.set_facecolor("#16213e")

    # Background dots
    ax.scatter(bg_x, bg_y, s=dot_size, c="#4a5568", alpha=0.5,
               edgecolors="none", marker="s",
               label=f"All boards ({len(boards)})")

    # Miss dots (behind ACBL)
    if miss_ids and miss_x:
        ax.scatter(miss_x, miss_y, s=dot_size * 1.5, c="#00ccff", alpha=0.9,
                   edgecolors="none", marker="s",
                   label=f"Hotspot Miss ({len(miss_x)})", zorder=3)

    # ACBL dots on top
    ax.scatter(acbl_x, acbl_y, s=dot_size * 1.5, c="#ff4444", alpha=0.9,
               edgecolors="none", marker="s",
               label=f"ACBL Disclosure ({len(acbl_x)})", zorder=4)

    # Year labels on left
    for y_pos, label in year_label_positions:
        ax.text(-8, y_pos, label, fontsize=11, color="white",
                ha="right", va="center", fontweight="bold")

    # Invert Y so earliest year is at top
    ax.invert_yaxis()

    ax.set_xlim(-15, cols_per_row + 5)
    ax.set_title(title, fontsize=14, color="white", pad=15)

    # Hide axes
    ax.set_xticks([])
    ax.set_yticks([])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)

    legend = ax.legend(loc="upper right", fontsize=10,
                       facecolor="#16213e", edgecolor="#4a5568",
                       labelcolor="white")

    fig.tight_layout()
    fig.savefig(output_path, dpi=200, facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate packed timeline charts from an EDGAR Defense workbook."
    )
    parser.add_argument("workbook", type=Path, help="Path to the EDGAR Defense .xlsx workbook")
    args = parser.parse_args()

    workbook = args.workbook
    if not workbook.exists():
        print(f"Error: {workbook} not found")
        return

    output_dir = workbook.parent
    case_name = workbook.stem  # e.g. "EDGAR Reynolds 2026-03-08"

    print(f"Loading {workbook.name}...")
    boards, acbl_ids, miss_ids = load_data(workbook)
    n_boards = len(boards)
    print(f"  {n_boards} boards, {len(acbl_ids)} ACBL, {len(miss_ids)} hotspot misses")

    # Chart 1: Small dots
    make_packed_plot(
        boards, acbl_ids,
        title=f"{n_boards:,} Boards — Each Dot Is One Hand",
        output_path=output_dir / "timeline_1_packed.pdf",
        dot_size=8,
    )

    # Chart 2: Larger dots
    make_packed_plot(
        boards, acbl_ids,
        title=f"{n_boards:,} Boards — ACBL Disclosure Hands in Red",
        output_path=output_dir / "timeline_2_packed_large.pdf",
        dot_size=20,
    )

    # Chart 3: With hotspot misses in blue
    make_packed_plot(
        boards, acbl_ids,
        title=f"{n_boards:,} Boards — ACBL (Red) + Hotspot Misses (Blue)",
        output_path=output_dir / "timeline_3_packed_misses.pdf",
        dot_size=8,
        miss_ids=miss_ids,
    )

    print(f"\nDone. PDFs saved to: {output_dir}")


if __name__ == "__main__":
    main()
